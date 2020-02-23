#! python3
# -*- encoding: utf-8 -*-

# TODO: TBU
# 指定のプルダウンリスト（option  valueとタグ間の文字列）を表示する。

import sys, os, requests, bs4, lxml
import urllib.robotparser
import datetime
import logging
from openpyxl import Workbook

#logging.disable(logging.DEBUG)
logging.basicConfig(filename='debug_{}.log'.format(__file__), level=logging.DEBUG,
        format=' %(asctime)s - %(levelname)s - %(message)s')
logging.debug('START: {}'.format(__file__))

USAGE = '''
使い方
対象URL、selectタグのid又はclass名をコマンドライン引数として渡す
test_soup.py argv[1](=URL) argv[2](=selector ('id or class'))

e.g.,
test_soup.py https://www.example.com/foo/bar/ #example_id
'''
def com_end():
    logging.debug('END {}'.format(__file__))
    print('処理を終了します')
    sys.exit()

if len(sys.argv) < 3:
    print(USAGE)
else:
    try:
        url = sys.argv[1]
        selector = sys.argv[2]
        res = requests.get(url)
        print(res)
        # TODO: encodingを確認
        logging.debug('res.encoding: {}'.format(res.encoding))
        logging.debug('res.apparent_encoding: {}'.format(res.apparent_encoding))
        #res.encoding = res.apparent_encoding
        res.raise_for_status()

        # robot.txtを確認
        rp = urllib.robotparser.RobotFileParser()
        rp.set_url(url)
        rp.read()
        if rp.can_fetch('*', url):
            ROBOT_TXT_OK = 'OK: Robot.txt allow UA to fetch the URL'
            logging.debug(ROBOT_TXT_OK)
            print(ROBOT_TXT_OK)
        else:
            ROBOT_TXT_NG = 'NG: Robot.txt DISALLOW UA to fetch the URL'
            logging.debug(ROBOT_TXT_NG)
            print(ROBOT_TXT_NG )
            com_end()

        # TODO: 非同期で生成されるオプションの場合

        soup = bs4.BeautifulSoup(res.text, 'lxml')
        # optionタグを取得
        link_elems = soup.select(' '.join([selector, 'option']))
        logging.debug('link_elems: {}'.format(link_elems))
        if len(link_elems) == 0:
            print('指定のオプションタグは見つかりませんでした')
        else:
            row = 1
            # 取得したオプションをファイルに出力する（xlsx）
            wb = Workbook()
            ws = wb.active
            ws.title = 'option_list'
            now = datetime.datetime.now()
            fmt_now = now.strftime('%y%m%d_%H%M%S')
            for i, v in enumerate(link_elems):
                cell_v, cell_s = 'A' + str(row), 'B' + str(row)
                ws[cell_v], ws[cell_s] = v['value'], v.string
                print('[value] : [str] - {} : {}'.format(v['value'], v.string))
                row += 1
            wb.save('opt_list_{}.xlsx'.format(fmt_now))
    except Exception as e:
        logging.debug('ERROR: {}'.formattype(e))
        print('エラーが発生しました：{}'.format(e))
        com_end()
    finally:
        com_end()
